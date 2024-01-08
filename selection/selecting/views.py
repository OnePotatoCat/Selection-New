from django.shortcuts import render, redirect
from django.http import HttpResponseRedirect, JsonResponse, Http404, HttpResponse, FileResponse
from django.urls import reverse
from django.contrib.staticfiles import finders
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, logout
from django.contrib import messages
from django.core.paginator import Paginator
from django.templatetags.static import static
from django.template import loader
from django.template.loader import render_to_string
from django import forms
from .models import Series, Unit, Compressor, Condenser, FlowOrientation, MotorType, Calculation, Cart, History

import json
import os
import zipfile
import re
from pyfluids import HumidAir, InputHumidAir
from reportlab.lib.pagesizes import letter, A4
from reportlab.pdfgen import canvas
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from openpyxl import load_workbook

from svglib.svglib import svg2rlg
from PIL import Image

# import Selection
import datetime
from . import Selection as sel
from . import SelectionCW as sel_cw
from .Compressor import Compressor_Cal as cond
from .Evaporator import Evaporator_Cal as evap
from .Condenser import Condenser_Cal
from .Fan import Fan_Cal
from .Unit import Unit_Cal

CONDENSER_MODEL = []

# Create your views here.
class NewTaskForm(forms.Form):
    temp = forms.FloatField(label="Return Air Temperature", min_value=5.0, max_value=40.0,
        widget=forms.NumberInput(attrs = {
            'placeholder': '24',
            'style': 'width:200px;',
            'class':'form-control'
        }))
    rh = forms.FloatField(label = "Return Air Relative Humidity", min_value=10, max_value=99.0,
        widget=forms.NumberInput(attrs = {
            'placeholder': '50',
            'style': 'width:200px;',
            'class':'form-control'
        }))
    airflow = forms.FloatField(label = "Airflow Rate, m3/hr", min_value=4000, max_value=8000,
        widget=forms.NumberInput(attrs = {
            'placeholder': '6650',
            'style': 'width:200px;',
            'class':'form-control'
        }))


class Calculation_Form(forms.Form):

    compressor = forms.ChoiceField(choices = [])
    fan = forms.ChoiceField(choices = [])
    condenser = forms.ChoiceField(choices = [])

    def __init__(self, compressor, fan, condensers):
        super(Calculation_Form, self).__init__()
        self.fields['compressor'] = forms.ChoiceField(choices = compressor, 
            widget = forms.Select(attrs={
            'style': 'width:200px;',
            'class': 'form-control'
        }))

        self.fields['fan'] = forms.ChoiceField(choices = fan, 
            widget = forms.Select(attrs={
            'style': 'width:200px;',
            'class': 'form-control'
        }))

        self.fields['condenser'] = forms.ChoiceField(choices = condensers, 
            widget = forms.Select(attrs={
            'style': 'width:200px;',
            'class': 'form-control'
        }))
        
    temp = forms.FloatField(label="Return Air Temperature", min_value=5.0, max_value=40.0,
        widget=forms.NumberInput(attrs = {
            'placeholder': '24',
            'style': 'width:200px;',
            'class':'form-control'
        }))

    rh = forms.FloatField(label = "Return Air Relative Humidity", min_value=10, max_value=99.0,
        widget=forms.NumberInput(attrs = {
            'placeholder': '50',
            'style': 'width:200px;',
            'class':'form-control'
        }))

    airflow = forms.FloatField(label = "Airflow Rate, m3/hr", min_value=4000, max_value=8000,
        widget=forms.NumberInput(attrs = {
            'placeholder': '6650',
            'style': 'width:200px;',
            'class':'form-control'
        }))

    esp = forms.FloatField(label = "External Static Pressure, Pa", min_value=10, max_value=400,
        widget=forms.NumberInput(attrs = {
            'placeholder': '50',
            'style': 'width:200px;',
            'class':'form-control'
        }))


class NewUnitSelectionForm(forms.Form):
    def __init__(self):
        super(NewUnitSelectionForm, self).__init__()
        units = Unit.objects.all()
        ids, models = units.values_list('id').order_by('id'), units.values_list('model').order_by('id')
        id_model_pairs = [(ids[i][0], models[i][0].upper()) for i in range(0, len(ids))]
        self.fields['selections'] = forms.ChoiceField(
            choices = [(pair[0], pair[1]) for pair in id_model_pairs])

date_time_format = "%Y %b %d %H:%M"
date_time_ymd = '%Y-%m-%d %H:%M'
date_format = "%Y %b %d"

# ------------------------------------ #
#          View Functions 
# ------------------------------------ #
def index(request):
    if not request.user.is_authenticated:
        return HttpResponseRedirect("/login")
    
    context = {
        "username" :request.user.username,
        "admin" : request.user.is_staff,
    }
    return render(request, "selecting/layout.html", context)


def show_series(request):
    if request.method =='GET':
        return HttpResponseRedirect("/selecting")
    
    # if request.method == 'POST':
    seriess = Series.objects.all().order_by('arrange_id')
    for series in seriess:
        series.static_url = "/staticfiles/selecting/series/" + series.series_name.lower()+".png"
    context ={"series" :seriess}
    template = loader.get_template("selecting/series_album.html")
    selection_html = template.render(context, request)
    return HttpResponse(selection_html)

        
def show_unit_selection(request, series):
    series_name = Series.objects.get(pk=int(series)).series_name
    if series_name == "CIDC":
        return render(request, "selecting/cidc.html")
    
    if series_name == "CIRC":
        return render(request, "selecting/circ.html")
    
    units = Unit.objects.filter(series = int(series)).order_by('arrange_id')
    if units.count() == 0:
        context = {"series":series_name}
        return render(request, "selecting/not_available.html", context)

    context = {
        "series": series_name,
        "units" : units, 
        "admin" : request.user.is_staff,
        }
    units_dict= {}
    for unit in units:
        units_dict[unit.id] = unit.model.upper()

    if series_name =="EC":
        template = loader.get_template("selecting/newselection_cw.html")
    else:
        template = loader.get_template("selecting/newselection.html")
        
    selection_html = template.render(context, request)
    return HttpResponse(selection_html)


def show_components(request, unit):
    data = {}
    unit = Unit.objects.get(pk=int(unit))
    
    # Flow orientation data
    flows = unit.flow_direction.all()
    flow_dict = {}
    for flow in flows:
        flow_dict[flow.id] = flow.flow_orientaion.upper()
    data["flow"] = flow_dict

    # Fan model data
    fan = unit.fan
    fan_dict = {}
    fan_dict[fan.id] = fan.model.upper()
    data["fan"] = fan_dict
    # Fan motor type data
    data["fan_motor_type"] = unit.fan.type

    # Default compressor speed
    data["default_comp_speed"] = unit.default_comp_speed

    # Default airflow
    data["default_airflow"] = unit.default_airflow

    # ----- DX ----- # 
    # Compressor model data
    if unit.compressor is not None:
        comp = unit.compressor
        comp_dict = {}
        comp_dict[comp.id] = comp.model.upper()
        data["compressor"] = comp_dict

    # Condenser model data
    if unit.condenser:
        condensers = unit.condenser.all()
        cond_dict = {}
        for condenser in condensers:
            cond_dict[condenser.id] = condenser.get_model_name()
        data["condenser"] = cond_dict

    # Evaporator airflow limit data
    if unit.evaporator:
        data["max_airflow"] = unit.evaporator.max_airflow
        data["min_airflow"] = unit.evaporator.min_airflow
    

    # ----- CW ----- # 
    # chillwater coil airflow limit data
    if unit.cw_coil:
        data["max_airflow"] = unit.cw_coil.max_airflow
        data["min_airflow"] = unit.cw_coil.min_airflow

    jsonData = json.dumps(data)
    return JsonResponse(data)


def set_default_airflow(request, unit):
    unit = Unit.objects.get(pk=int(unit))
    return JsonResponse({"airflow": unit.default_airflow})


def inverter_compressor(request, comp):
    compressor = Compressor.objects.get(pk=int(comp))
    
    return JsonResponse({"inverter": compressor.inverter and request.user.is_staff})


def ac_fan_airflow(unit, esp, filter):
    unit_obj = Unit.objects.get(pk=int(unit))
    unit_cal = Unit_Cal(unit, unit_obj.evaporator.id, 1, 1, 1, unit_obj.fan.id, esp, filter)
    fan = Fan_Cal(unit_obj.fan.id)

    temp_airflow = unit_obj.default_airflow
    max_airflow = unit_obj.evaporator.max_airflow
    min_airflow = unit_obj.evaporator.min_airflow
    i = 1
    while i < 500:
        tsp = unit_cal.get_tsp(temp_airflow)
        sp_diff = tsp - fan.get_ac_staticpressure(temp_airflow)
        if abs(sp_diff) < 0.5:
            break
        if sp_diff > 0:
            max_airflow = temp_airflow
        else:
            min_airflow = temp_airflow

        temp_airflow = 0.5*(max_airflow + min_airflow)
        i += 1

    return temp_airflow


def ac_fan_esp(request, unit, airflow, filter):
    return JsonResponse(ac_fan_esp_internal(unit, airflow, filter))


def ac_fan_esp_internal(unit, airflow, filter):
    max_esp = 200
    min_esp = 20
    unit_obj = Unit.objects.get(pk=int(unit))
    unit_cal = Unit_Cal(unit, unit_obj.evaporator.id, 1, 1, 1, unit_obj.fan.id, 0, filter)
    fan = Fan_Cal(unit_obj.fan.id)

    fan_staticpressure = fan.get_ac_staticpressure(airflow)
    unit_isp = unit_cal.get_isp(airflow)
    esp =  fan_staticpressure - unit_isp

    if esp < min_esp:
        esp = 20
        airflow = ac_fan_airflow(unit, esp, filter)

    elif max_esp < esp:
        esp = 200
        airflow = ac_fan_airflow(unit, esp, filter)
    
    return ({"airflow": round(airflow), "esp": round(esp)})


def calculate_capacity(request):
    if request.method =="POST":
        form = request.POST
        unit_id = int(form["unit"])
        flow_id = int(form["flow"])
        fan_id = int(form["fan"])
        inlet_temp = float(form["temp"])
        rh = float(form["rh"])

        airflow = float(form["airflow"])
        filter_type = form["filter"].lower()
        esp = float(form["esp"])

        fan = Fan_Cal(fan_id)
        if fan.type==1:
            dict = ac_fan_esp_internal(unit_id, airflow, filter_type)
            airflow = dict["airflow"]
            esp = dict["esp"]

        # DX
        if form['type']=="DX":
            cond_id = int(form["cond"])
            evap_id = Unit.objects.get(pk=int(unit_id)).evaporator.id
            comp_id = int(form["comp"])
            amb_temp = float(form["amb_temp"])

            if (form["comp_sp"] != ''):
                comp_speed = float(form["comp_sp"])
            else:
                comp_speed = float(0)

            result = sel.main(unit_id, evap_id, cond_id, comp_id, fan_id, inlet_temp, rh, airflow, esp, amb_temp, comp_speed, filter_type)
       
        # CW
        elif form['type']=="CW":
            # TODO:
            # vavle_id = 0
            cw_id=Unit.objects.get(pk=unit_id).cw_coil.id
            t_water_inlet = float(form['temp_water_in'])
            t_water_outlet = float(form['temp_water_out'])
            result = sel_cw.main(unit_id, cw_id, fan_id, inlet_temp, rh, airflow, t_water_inlet, t_water_outlet, esp, filter_type)
            return

        # result = sel.main(unit_id, evap_id, cond_id, comp_id, fan_id, inlet_temp, rh, airflow, esp, amb_temp, comp_speed, filter_type)
        result["calculation id"] = ""

        if result["converged"]:
            new_calculation = Calculation(
                add_to_cart = False,
                date_time = datetime.datetime.now().strftime(date_time_ymd),
                model = Unit.objects.get(pk=int(unit_id)),
                flow_orientaion = FlowOrientation.objects.get(pk=int(flow_id)),
                cond = Condenser.objects.get(pk=int(cond_id)),
                inlet_temp = inlet_temp,
                inlet_rh = rh,
                airflow = airflow,
                esp = esp,
                amb_temp = amb_temp,
                filter = filter_type,
                comp = Compressor.objects.get(pk=int(comp_id)),
                comp_spd = comp_speed,
                total_cap = result["capacity"]["Total Capacity"][0],
                sen_cap = result["capacity"]["Total Sensible Cap."][0],
                heat_rejection = round(result["capacity"]["Total Capacity"][0]/Unit.objects.get(pk=int(unit_id)).number_of_compressor + result["compressor"]["Comp. Power"][0], 2),
                fan_power = result["fan"]["Fan Power"][0],
                fan_rpm = result["fan"]["Fan RPM"][0],
                tsp = result["fan"]["Total Static Pressure"][0],
                t_evap = result["compressor"]["Evaporating Temp."][0],
                t_cond = result["compressor"]["Condensing Temp."][0],
                off_temp = result["air"]["Off Coil Temperature"][0],
                off_rh = result["air"]["Off Coil RH"][0],
                outlet_temp = result["air"]["Outlet Temperature"][0], 
                outlet_rh = result["air"]["Outlet RH"][0],
            )
            new_calculation.save()
            
            result["calculation id"] = new_calculation.id
        return JsonResponse(result)
    

def download_pdf(request, pdf):
    pdf = os.path.join("selecting/static/selecting/pdf", str(pdf))
    response = FileResponse(open(pdf, 'rb'), content_type='application/pdf', as_attachment=True)
    response['Content-Disposition'] = 'attachment; filename = "catalogue_pdfs.pdf'
    response['Content-Length'] = os.path.getsize(pdf)
    return response


def add_calculation_to_cart(request, cal_id):
    calculation_data = Calculation.objects.get(pk=int(cal_id))
    calculation_data.add_to_cart = True
    cart = Cart(
        user = User.objects.get(pk=int(request.user.id)),
        calculation = Calculation.objects.get(pk=int(cal_id))
    )
    cart.save()
    
    return JsonResponse({"success": calculation_data.add_to_cart,
                         "message": "Added to cart successfully" })


def show_cart(request):
    if request.method =='GET':
        return HttpResponseRedirect("/selecting")

    user_cart = Cart.objects.filter(user = request.user.id)
    cart_dict = {}

    for item in user_cart:
        net_total_cap = round(item.calculation.total_cap - item.calculation.fan_power, 1)
        net_sen_cap = round(item.calculation.sen_cap - item.calculation.fan_power, 1)

        cart_dict[str(item.id)] = {
            "model"         : f"{item.calculation.model} - {item.calculation.cond}",
            "total_cap"     : f"{round(item.calculation.total_cap,1)} kW",
            "sen_cap"       : f"{round(item.calculation.sen_cap,1)} kW",
            "net_total_cap" : f"{net_total_cap} kW",
            "net_sen_cap"   : f"{net_sen_cap} kW",
            "airflow"       : f"{item.calculation.airflow} m3/hr",
            "inlet_temp"    : f"{item.calculation.inlet_temp} °C", 
            "inlet_rh"      : f"{item.calculation.inlet_rh} %",
            "outlet_temp"   : f"{item.calculation.outlet_temp} °C", 
            "outlet_rh"     : f"{item.calculation.outlet_rh} %",
            "filter"        : f"{item.calculation.filter.upper()}",
            "datetime"      : f"{item.calculation.date_time.strftime(date_time_format)}"
        }

    context = {
        "cart": cart_dict, 
        "admin": request.user.is_staff,
        }
    
    template = loader.get_template("selecting/cart.html")
    cart_html = template.render(context, request)
    return HttpResponse(cart_html)


# Generate reports in pdf/excel and compressed to zip file for downloading
#   pdf/excel files are generated in    "/repots/pdf/*account id*"
#   zip files are generated in "/repots/zip/*account id*""
#   files are removed in the respective directories before generating new ones
#   then zip file is generated for download
def generate_reports(request, cal_ids):
    pdfs_directory = os.path.join("selecting/reports/pdf", str(request.user.id))
    os.makedirs(pdfs_directory, exist_ok=True)
    for filename in os.listdir(pdfs_directory):
        file_path = os.path.join(pdfs_directory, filename)
        os.remove(file_path)

    zip_directory = os.path.join("selecting/reports/zip", str(request.user.id))
    os.makedirs(zip_directory, exist_ok=True)
    for filename in os.listdir(zip_directory):
        file_path = os.path.join(zip_directory, filename)
        os.remove(file_path)
    
    ids = cal_ids.split(",")
    for id in ids:
        cart = Cart.objects.get(pk=int(id))
        filename = f'{cart.id}_{cart.calculation.model}-{cart.calculation.cond}'
        file_path = os.path.join(pdfs_directory, filename)
        history = add_to_history(cart, "Gen")
        generate_pdf(f'{file_path}.pdf', history)
        generate_excel(f'{file_path}.xlsx', history)
        cart.delete()

    zip_file_path = os.path.join(zip_directory, "generated_pdfs.zip")
    with zipfile.ZipFile(zip_file_path, 'w') as zip_file:
        for filename in os.listdir(pdfs_directory):
            file_path = os.path.join(pdfs_directory, filename)
            zip_file.write(file_path, arcname = filename)
    
    response = FileResponse(open(zip_file_path, 'rb'), content_type='application/zip', as_attachment=True)
    response['Content-Disposition'] = 'attachment; filename = "generated_pdfs.zip'
    response['Content-Length'] = os.path.getsize(zip_file_path)
    return response


def delete_cart_items(request, cal_ids):
    ids = cal_ids.split(",")
    for id in ids:
        cart = Cart.objects.get(pk=int(id))
        history = add_to_history(cart, "Del")
        cart.delete()
    return show_cart(request)


def add_to_history(cart, status):
    history = History(
            user = User.objects.get(pk=int(cart.user.id)),
            calculation = Calculation.objects.get(pk=int(cart.calculation.id)),
            generated_date_time = datetime.datetime.now().strftime(date_time_ymd),
            status = status
        )
    history.save()
    return history


def show_history(request):
    if request.method =='GET':
        return HttpResponseRedirect("/selecting")

    history_list = History.objects.filter(user = request.user.id).order_by('generated_date_time').reverse()
    paginator = Paginator(history_list, 10)
    page_number = request.GET.get('page')
    try:
        page_obj = paginator.get_page(page_number)
    except:
        page_number = 1
        page_obj = paginator.get_page(page_number)
    context = {'page_obj': page_obj,
               'page_list': list(paginator.get_elided_page_range(number = page_number, on_each_side=1, on_ends=2))}

    template = loader.get_template("selecting/history.html")
    historyt_html = template.render(context, request)
    return HttpResponse(historyt_html)


def generate_pdf(file_path, history):
    pdfmetrics.registerFont(TTFont('Vera', 'Vera.ttf'))
    pdfmetrics.registerFont(TTFont('VeraBd', 'VeraBd.ttf'))
    hist = History.objects.get(pk=int(history.id))

    pdf = canvas.Canvas(file_path, pagesize=A4)
    width, height = A4
    padding = 20
    starting_spacer = 25
    subtitle_spacer = 40
    line_spacer = 5
    row_spacer = 15

    def draw_subtitle(text):
        pdf.setFont('VeraBd', 18)
        nonlocal current_height
        current_height -= subtitle_spacer
        pdf.drawString(padding, current_height, text)
        current_height -= line_spacer
        pdf.line(padding, current_height, width-padding*2, current_height)

    def draw_row_entry(label, value):
        pdf.setFont('VeraBd', 11)
        nonlocal current_height
        current_height -= row_spacer
        pdf.drawString(padding*2, current_height, label)
        pdf.drawString(width/2, current_height, value)


    title = 'Product Selection Sheet'
    date_time = history.generated_date_time
    temp_model = re.compile("([a-zA-Z]+)([0-9]+)([a-zA-Z]+[0-9])")
    prefix, cap, suffix = temp_model.match(hist.calculation.model.model).groups()
    # print(suffix)
    full_model = f"{prefix}{hist.calculation.flow_orientaion.discharge_orientation}{cap}{suffix}"
    # print(full_model)

    # Logo and top bar
    bar_height = 45
    logo_path = finders.find('selecting/logo_report.png')
    logo = Image.open(logo_path)
    logo_width, logo_height = logo.size
    logo_scaling = (bar_height-5)/logo_height
    pdf.setFillColorRGB(0,0,0)
    pdf.rect(0, height-bar_height, width, bar_height, fill=True, stroke=False)
    pdf.drawImage(logo_path, 5, height-logo_height*logo_scaling, logo_width*logo_scaling, logo_height*logo_scaling)

    # IDs
    pdf.setFont('VeraBd', 9)
    pdf.setFillColorRGB(1,1,1) 
    pdf.drawString(width-165, height-bar_height+25, "User ID")
    pdf.drawString(width-120, height-bar_height+25, f": {history.user.id}")
    pdf.drawString(width-165, height-bar_height+15, "Cal. ID")
    pdf.drawString(width-120, height-bar_height+15, f": {history.id}")
    pdf.drawString(width-165, height-bar_height+5, "Date")
    pdf.drawString(width-120, height-bar_height+5, f": {date_time}")

    # Reset text color to black, start height
    pdf.setFillColorRGB(0,0,0) 
    current_height = height-bar_height-starting_spacer+subtitle_spacer

    # Subtitle - Product Information
    draw_subtitle("Product Information")
    # Series Name
    draw_row_entry("Series", f"{hist.calculation.model.series}")
    # Model Name
    draw_row_entry("Model", f"{full_model}")
    # Flow configuration
    draw_row_entry("Flow Configuration", f"{hist.calculation.flow_orientaion}")
    # Refrigerant
    draw_row_entry("Refrigerant", f"{hist.calculation.comp.refrigerant}")
    # Unit Dimension
    draw_row_entry("Unit Dimension LxDxH", f"{hist.calculation.model.length} x {hist.calculation.model.depth} x {hist.calculation.model.height} mm\u00b3")
    # Power Supply
    draw_row_entry("Power Supply", f"{hist.calculation.model.power_supply}")

    # Subtitle - Fan Information
    draw_subtitle("Fan Information")
    # Fan Type
    draw_row_entry("Fan Type", f"{MotorType.objects.get(pk=hist.calculation.model.fan.type).type} {hist.calculation.model.fan.size} D")
    # No. of Fan
    draw_row_entry("No. of Fan", f"{hist.calculation.model.number_of_fan}")
    # Fan Speed
    draw_row_entry("Fan Speed", f"{hist.calculation.fan_rpm} RPM")
    # Filter Type
    draw_row_entry("Filter Type", f"{hist.calculation.filter.upper()}")
    # Airflow Rate
    draw_row_entry("Airflow Rate", f"{hist.calculation.airflow} m\u00b3/hr")
    # ESP
    draw_row_entry("External Static Pressure", f"{hist.calculation.esp} Pa")
    # TSP
    draw_row_entry("Total Static Pressure", f"{hist.calculation.tsp} Pa")
    # Motor Heat
    draw_row_entry("Motor Heat", f"{hist.calculation.fan_power} kW")

    # Subtitle - Entering Air Properties
    draw_subtitle("Entering Air Properties")
    # Dry Bulb Temperature
    db_inlet = hist.calculation.inlet_temp
    draw_row_entry("Dry Bulb Temperature", f"{db_inlet} °C")
    # RH
    rh_inlet = hist.calculation.inlet_rh
    draw_row_entry("Relative Humidity", f"{rh_inlet} %")
    # WB
    inlet_air = HumidAir().with_state(
        InputHumidAir.altitude(0),
        InputHumidAir.temperature(db_inlet),
        InputHumidAir.relative_humidity(rh_inlet)
    )
    wb_inlet = round(inlet_air.wet_bulb_temperature,1)
    draw_row_entry("Wet Bulb Temperature", f"{wb_inlet} °C")


    # Subtitle - Outlet Air Properties
    draw_subtitle("Outlet Air Properties")
    # Dry Bulb Temperature
    db_outlet = hist.calculation.outlet_temp
    draw_row_entry("Dry Bulb Temperature", f"{db_outlet} °C")
    # RH
    rh_outlet = hist.calculation.outlet_rh
    draw_row_entry("Relative Humidity", f"{rh_outlet} %")
    # WB
    outlet_air = HumidAir().with_state(
        InputHumidAir.altitude(0),
        InputHumidAir.temperature(db_outlet),
        InputHumidAir.relative_humidity(rh_outlet)
    )
    wb_outlet = round(outlet_air.wet_bulb_temperature,1)
    draw_row_entry("Wet Bulb Temperature", f"{wb_outlet} °C")


    # Subtitle - Refrigeration System/Circuit
    draw_subtitle("Refrigeration System/Circuit")
    # Ambient Temperature
    draw_row_entry("Ambient Temperature", f"{hist.calculation.amb_temp} °C")
    # No of Circuit
    draw_row_entry("No. of Circuit", f"{hist.calculation.model.number_of_compressor}")
    # Condenser Model
    draw_row_entry("Condenser Model", f"{hist.calculation.cond.get_model_name()}")
    # Heat Rejection
    draw_row_entry("Heat Rejection", f"{hist.calculation.heat_rejection} kW")
    # Compressor
    draw_row_entry("Compressor Size", f"{hist.calculation.model.compressor.hp}HP")
    # Evaporating Temperature
    draw_row_entry("Evaporating Temperature", f"{hist.calculation.t_evap} °C")
    # Condensing Temperature
    draw_row_entry("Condensing Temperature", f"{hist.calculation.t_cond} °C")
    # Compressor Power Input
    draw_row_entry("Compressor Power Input", f"{round(hist.calculation.heat_rejection - hist.calculation.total_cap/hist.calculation.model.number_of_compressor, 2)} kW")

    # Subtitle - Unit Capacity
    draw_subtitle("Unit Capacity")
    # Gross Total Capacity
    draw_row_entry("Gross Total Capacity", f"{hist.calculation.total_cap} kW")
    # Gross Sensible Capacity
    draw_row_entry("Gross Sensible Capacity", f"{hist.calculation.sen_cap} kW")
    # Gross SHR
    draw_row_entry("Gross SHR", f"{round(hist.calculation.sen_cap/hist.calculation.total_cap, 2)}")
    # Net Total Capacity
    draw_row_entry("Net Total Capacity", f"{round(hist.calculation.total_cap - hist.calculation.fan_power, 2)} kW")
    # Gross Sensible Capacity
    draw_row_entry("Net Sensible Capacity", f"{round(hist.calculation.sen_cap- hist.calculation.fan_power, 2)} kW")
    # Gross SHR
    draw_row_entry("Net SHR", f"{round((hist.calculation.sen_cap - hist.calculation.fan_power)/(hist.calculation.total_cap- hist.calculation.fan_power), 2)}")

    pdf.save()


def generate_excel(file_path, history):
    hist = History.objects.get(pk=int(history.id))
    date_time = history.generated_date_time
    temp_model = re.compile("([a-zA-Z]+)([0-9]+)([a-zA-Z]+[0-9])")
    prefix, cap, suffix = temp_model.match(hist.calculation.model.model).groups()
    full_model = f"{prefix}{hist.calculation.flow_orientaion.discharge_orientation}{cap}{suffix}"

    template_file = "selecting/reports/report_template.xlsx"
    xls = load_workbook(filename=template_file)
    sheet = xls.active

    # IDs
    sheet.cell(row=2, column=7).value = f"{history.user.id}"
    sheet.cell(row=3, column=7).value = f"{history.id}"
    sheet.cell(row=4, column=7).value = f"{date_time}"

    # Subtitle - Product Information
    # Series Name
    sheet.cell(row=7, column=4).value = f"{hist.calculation.model.series}"
    # Model Name
    sheet.cell(row=8, column=4).value = f"{full_model}"
    # Flow configuration
    sheet.cell(row=9, column=4).value = f"{hist.calculation.flow_orientaion}"
    # Refrigerant
    sheet.cell(row=10, column=4).value =  f"{hist.calculation.comp.refrigerant}"
    # Unit Dimension
    sheet.cell(row=11, column=4).value = f"{hist.calculation.model.length} x {hist.calculation.model.depth} x {hist.calculation.model.height} mm\u00b3"
    # Power Supply
    sheet.cell(row=12, column=4).value = f"{hist.calculation.model.power_supply}"

    # Subtitle - Fan Information
    # Fan Type
    sheet.cell(row=15, column=4).value = f"{MotorType.objects.get(pk=hist.calculation.model.fan.type).type} {hist.calculation.model.fan.size} D"
    # No. of Fan
    sheet.cell(row=16, column=4).value = f"{hist.calculation.model.number_of_fan}"
    # Fan Speed
    sheet.cell(row=17, column=4).value = f"{hist.calculation.fan_rpm} RPM"
    # Filter Type
    sheet.cell(row=18, column=4).value = f"{hist.calculation.filter.upper()}"
    # Airflow Rate
    sheet.cell(row=19, column=4).value = f"{hist.calculation.airflow} m\u00b3/hr"
    # ESP
    sheet.cell(row=20, column=4).value = f"{hist.calculation.esp} Pa"
    # TSP
    sheet.cell(row=21, column=4).value = f"{hist.calculation.tsp} Pa"
    # Motor Heat
    sheet.cell(row=22, column=4).value = f"{hist.calculation.fan_power} kW"

    # Subtitle - Entering Air Properties
    # Dry Bulb Temperature
    db_inlet = hist.calculation.inlet_temp
    sheet.cell(row=25, column=4).value = f"{db_inlet} °C"
    # RH
    rh_inlet = hist.calculation.inlet_rh
    sheet.cell(row=26, column=4).value = f"{rh_inlet} %"
    # WB
    inlet_air = HumidAir().with_state(
        InputHumidAir.altitude(0),
        InputHumidAir.temperature(db_inlet),
        InputHumidAir.relative_humidity(rh_inlet)
    )
    wb_inlet = round(inlet_air.wet_bulb_temperature,1)
    sheet.cell(row=27, column=4).value = f"{wb_inlet} °C"

    # Subtitle - Outlet Air Properties
    # Dry Bulb Temperature
    db_outlet = hist.calculation.outlet_temp
    sheet.cell(row=30, column=4).value = f"{db_outlet} °C"
    # RH
    rh_outlet = hist.calculation.outlet_rh
    sheet.cell(row=31, column=4).value = f"{rh_outlet} %"
    # WB
    outlet_air = HumidAir().with_state(
        InputHumidAir.altitude(0),
        InputHumidAir.temperature(db_outlet),
        InputHumidAir.relative_humidity(rh_outlet)
    )
    wb_outlet = round(outlet_air.wet_bulb_temperature,1)
    sheet.cell(row=32, column=4).value = f"{wb_outlet} °C"

    # Subtitle - Refrigeration System/Circuit
    # Ambient Temperature
    sheet.cell(row=35, column=4).value = f"{hist.calculation.amb_temp} °C"
    # No of Circuit
    sheet.cell(row=36, column=4).value = f"{hist.calculation.model.number_of_compressor}"
    # Condenser Model
    sheet.cell(row=37, column=4).value = f"{hist.calculation.cond.get_model_name()}"
    # Heat Rejection
    sheet.cell(row=38, column=4).value = f"{hist.calculation.heat_rejection} kW"
    # Compressor
    sheet.cell(row=39, column=4).value = f"{hist.calculation.model.compressor.hp}HP"
    # Evaporating Temperature
    sheet.cell(row=40, column=4).value = f"{hist.calculation.t_evap} °C"
    # Condensing Temperature
    sheet.cell(row=41, column=4).value = f"{hist.calculation.t_cond} °C"
    # Compressor Power Input
    sheet.cell(row=41, column=4).value = f"{round(hist.calculation.heat_rejection - hist.calculation.total_cap/hist.calculation.model.number_of_compressor, 2)} kW"

    # Subtitle - Unit Capacity
    # Gross Total Capacity
    sheet.cell(row=45, column=4).value = f"{hist.calculation.total_cap} kW"
    # Gross Sensible Capacity
    sheet.cell(row=46, column=4).value = f"{hist.calculation.sen_cap} kW"
    # Gross SHR
    sheet.cell(row=47, column=4).value = f"{round(hist.calculation.sen_cap/hist.calculation.total_cap, 2)}"
    # Net Total Capacity
    sheet.cell(row=48, column=4).value = f"{round(hist.calculation.total_cap - hist.calculation.fan_power, 2)} kW"
    # Gross Sensible Capacity
    sheet.cell(row=49, column=4).value = f"{round(hist.calculation.sen_cap- hist.calculation.fan_power, 2)} kW"
    # Gross SHR
    sheet.cell(row=50, column=4).value = f"{round((hist.calculation.sen_cap - hist.calculation.fan_power)/(hist.calculation.total_cap- hist.calculation.fan_power), 2)}"

    xls.save(file_path) 